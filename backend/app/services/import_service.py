"""Service for importing data from CSV/Excel files."""

import base64
import contextlib
import csv
import io
import logging
import re
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation
from email.utils import parseaddr
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BadRequestException, NotFoundException
from app.models.client_profile import ClientProfile
from app.models.enums import TaskPriority, TaskStatus
from app.models.partner import PartnerProfile
from app.models.program import Program
from app.models.user import User
from app.schemas.import_schemas import (
    ColumnMapping,
    ImportEntityType,
    ImportError,
    ImportFieldDefinition,
    ImportJobResponse,
    ImportStatus,
    ImportTemplateResponse,
    ImportWarning,
)
from app.services.client_service import client_service
from app.services.duplicate_detection_service import check_duplicates

logger = logging.getLogger(__name__)

# In-memory storage for import jobs (in production, use Redis or database)
_import_jobs: dict[str, dict[str, Any]] = {}


# Field definitions for each entity type
CLIENT_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="legal_name",
        display_name="Legal Name",
        description="Client's legal name",
        required=True,
        field_type="string",
        example_values=["John Smith", "Acme Corporation"],
    ),
    ImportFieldDefinition(
        name="display_name",
        display_name="Display Name",
        description="Preferred name or nickname",
        required=False,
        field_type="string",
        example_values=["John", "Acme Corp"],
    ),
    ImportFieldDefinition(
        name="entity_type",
        display_name="Entity Type",
        description="Type of entity",
        required=False,
        field_type="enum",
        enum_values=["individual", "family_office", "trust", "corporation"],
        example_values=["individual", "corporation"],
    ),
    ImportFieldDefinition(
        name="jurisdiction",
        display_name="Jurisdiction",
        description="Primary jurisdiction",
        required=False,
        field_type="string",
        example_values=["United States", "United Kingdom", "Singapore"],
    ),
    ImportFieldDefinition(
        name="tax_id",
        display_name="Tax ID",
        description="Tax identification number",
        required=False,
        field_type="string",
        example_values=["123-45-6789"],
    ),
    ImportFieldDefinition(
        name="primary_email",
        display_name="Primary Email",
        description="Primary contact email",
        required=True,
        field_type="email",
        example_values=["john@example.com"],
    ),
    ImportFieldDefinition(
        name="secondary_email",
        display_name="Secondary Email",
        description="Secondary contact email",
        required=False,
        field_type="email",
        example_values=["john.d@example.com"],
    ),
    ImportFieldDefinition(
        name="phone",
        display_name="Phone",
        description="Contact phone number",
        required=False,
        field_type="phone",
        example_values=["+1 (555) 123-4567"],
    ),
    ImportFieldDefinition(
        name="address",
        display_name="Address",
        description="Physical address",
        required=False,
        field_type="string",
        example_values=["123 Main St, New York, NY 10001"],
    ),
    ImportFieldDefinition(
        name="communication_preference",
        display_name="Communication Preference",
        description="Preferred communication channel",
        required=False,
        field_type="enum",
        enum_values=["email", "phone", "sms", "whatsapp"],
        example_values=["email"],
    ),
    ImportFieldDefinition(
        name="sensitivities",
        display_name="Sensitivities",
        description="Any sensitivities to be aware of",
        required=False,
        field_type="string",
        example_values=["Prefers evening calls"],
    ),
    ImportFieldDefinition(
        name="special_instructions",
        display_name="Special Instructions",
        description="Special handling instructions",
        required=False,
        field_type="string",
        example_values=["Contact assistant first"],
    ),
    ImportFieldDefinition(
        name="birth_date",
        display_name="Birth Date",
        description="Date of birth (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["1980-01-15"],
    ),
    ImportFieldDefinition(
        name="assigned_rm_email",
        display_name="Assigned RM Email",
        description="Email of the relationship manager to assign",
        required=False,
        field_type="email",
        example_values=["rm@example.com"],
    ),
]

PARTNER_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="firm_name",
        display_name="Firm Name",
        description="Name of the partner firm",
        required=True,
        field_type="string",
        example_values=["Global Advisory Partners"],
    ),
    ImportFieldDefinition(
        name="contact_name",
        display_name="Contact Name",
        description="Primary contact person's name",
        required=True,
        field_type="string",
        example_values=["Jane Doe"],
    ),
    ImportFieldDefinition(
        name="contact_email",
        display_name="Contact Email",
        description="Primary contact email",
        required=True,
        field_type="email",
        example_values=["jane@globaladvisory.com"],
    ),
    ImportFieldDefinition(
        name="contact_phone",
        display_name="Contact Phone",
        description="Contact phone number",
        required=False,
        field_type="phone",
        example_values=["+44 20 1234 5678"],
    ),
    ImportFieldDefinition(
        name="capabilities",
        display_name="Capabilities",
        description="Comma-separated list of capabilities",
        required=False,
        field_type="string",
        example_values=["Legal, Tax Advisory, Immigration"],
    ),
    ImportFieldDefinition(
        name="geographies",
        display_name="Geographies",
        description="Comma-separated list of geographies served",
        required=False,
        field_type="string",
        example_values=["United States, Europe, Asia"],
    ),
    ImportFieldDefinition(
        name="notes",
        display_name="Notes",
        description="Additional notes about the partner",
        required=False,
        field_type="string",
        example_values=["Preferred partner for tax matters"],
    ),
]

PROGRAM_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="title",
        display_name="Program Title",
        description="Title of the program",
        required=True,
        field_type="string",
        example_values=["Estate Planning 2024"],
    ),
    ImportFieldDefinition(
        name="client_email",
        display_name="Client Email",
        description="Email of the client (must exist)",
        required=True,
        field_type="email",
        example_values=["client@example.com"],
    ),
    ImportFieldDefinition(
        name="objectives",
        display_name="Objectives",
        description="Program objectives",
        required=False,
        field_type="string",
        example_values=["Set up trust structure"],
    ),
    ImportFieldDefinition(
        name="scope",
        display_name="Scope",
        description="Program scope",
        required=False,
        field_type="string",
        example_values=["US and UK jurisdictions"],
    ),
    ImportFieldDefinition(
        name="budget_envelope",
        display_name="Budget Envelope",
        description="Budget amount",
        required=False,
        field_type="number",
        example_values=["50000", "100000.00"],
    ),
    ImportFieldDefinition(
        name="start_date",
        display_name="Start Date",
        description="Program start date (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["2024-01-01"],
    ),
    ImportFieldDefinition(
        name="end_date",
        display_name="End Date",
        description="Program end date (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["2024-12-31"],
    ),
    ImportFieldDefinition(
        name="status",
        display_name="Status",
        description="Program status",
        required=False,
        field_type="enum",
        enum_values=["planning", "active", "on_hold", "completed", "cancelled"],
        example_values=["planning", "active"],
    ),
]

TASK_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="title",
        display_name="Task Title",
        description="Title of the task",
        required=True,
        field_type="string",
        example_values=["Review legal documents"],
    ),
    ImportFieldDefinition(
        name="program_title",
        display_name="Program Title",
        description="Title of the associated program",
        required=False,
        field_type="string",
        example_values=["Estate Planning 2024"],
    ),
    ImportFieldDefinition(
        name="milestone_title",
        display_name="Milestone Title",
        description="Title of the associated milestone",
        required=False,
        field_type="string",
        example_values=["Phase 1: Research"],
    ),
    ImportFieldDefinition(
        name="description",
        display_name="Description",
        description="Task description",
        required=False,
        field_type="string",
        example_values=["Review all legal documents for compliance"],
    ),
    ImportFieldDefinition(
        name="due_date",
        display_name="Due Date",
        description="Task due date (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["2024-03-15"],
    ),
    ImportFieldDefinition(
        name="assigned_to_email",
        display_name="Assigned To Email",
        description="Email of the user to assign",
        required=False,
        field_type="email",
        example_values=["coordinator@example.com"],
    ),
    ImportFieldDefinition(
        name="priority",
        display_name="Priority",
        description="Task priority",
        required=False,
        field_type="enum",
        enum_values=["low", "medium", "high", "urgent"],
        example_values=["medium", "high"],
    ),
    ImportFieldDefinition(
        name="status",
        display_name="Status",
        description="Task status",
        required=False,
        field_type="enum",
        enum_values=["pending", "in_progress", "completed", "blocked"],
        example_values=["pending", "in_progress"],
    ),
]

FIELD_DEFINITIONS: dict[ImportEntityType, list[ImportFieldDefinition]] = {
    ImportEntityType.CLIENTS: CLIENT_FIELDS,
    ImportEntityType.PARTNERS: PARTNER_FIELDS,
    ImportEntityType.PROGRAMS: PROGRAM_FIELDS,
    ImportEntityType.TASKS: TASK_FIELDS,
}


def _normalize_column_name(name: str) -> str:
    """Normalize column name for matching."""
    # Lowercase, replace underscores/spaces with underscores, remove special chars
    normalized = name.lower().strip()
    normalized = re.sub(r"[\s\-]+", "_", normalized)
    normalized = re.sub(r"[^a-z0-9_]", "", normalized)
    return normalized


def _auto_detect_mappings(columns: list[str], entity_type: ImportEntityType) -> dict[str, str]:
    """Auto-detect column to field mappings based on column names."""
    field_defs = FIELD_DEFINITIONS.get(entity_type, [])

    # Build a lookup for field name variations
    field_variations: dict[str, str] = {}
    for field in field_defs:
        # Add field name
        field_variations[_normalize_column_name(field.name)] = field.name
        # Add display name
        field_variations[_normalize_column_name(field.display_name)] = field.name
        # Add common variations
        if field.name == "primary_email":
            field_variations["email"] = field.name
            field_variations["email_address"] = field.name
        if field.name == "legal_name":
            field_variations["name"] = field.name
            field_variations["client_name"] = field.name
        if field.name == "firm_name":
            field_variations["name"] = field.name
            field_variations["partner_name"] = field.name
        if field.name == "contact_email":
            field_variations["email"] = field.name
        if field.name == "contact_name":
            field_variations["name"] = field.name

    mappings: dict[str, str] = {}
    for col in columns:
        normalized = _normalize_column_name(col)
        if normalized in field_variations:
            mappings[col] = field_variations[normalized]

    return mappings


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse CSV content and return headers and rows."""
    text = content.decode("utf-8-sig")  # Handle BOM
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise BadRequestException("CSV file has no headers")
    columns = list(reader.fieldnames)
    rows = list(reader)
    return columns, rows


def _parse_excel(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse Excel content and return headers and rows."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise BadRequestException("Excel file has no sheets")

    # Get headers from first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        value = cell.value
        if value is not None:
            headers.append(str(value))
        else:
            headers.append(f"Column_{cell.column_letter}")

    # Get data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            row_dict[header] = str(value) if value is not None else ""
        rows.append(row_dict)

    wb.close()
    return headers, rows


def _validate_email(value: str) -> bool:
    """Validate email format."""
    _, email = parseaddr(value)
    return bool(email and "@" in email and "." in email.split("@")[1])


def _validate_phone(value: str) -> bool:
    """Validate phone format (basic check)."""
    # Allow digits, spaces, dashes, parentheses, plus
    cleaned = re.sub(r"[\d\s\-\(\)\+]", "", value)
    return len(cleaned) == 0 and len(re.sub(r"\D", "", value)) >= 7


def _validate_date(value: str) -> tuple[bool, str | None]:
    """Validate and normalize date format."""
    # Try common date formats
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"]
    for fmt in formats:
        try:
            from datetime import datetime as dt

            parsed = dt.strptime(value, fmt)
            return True, parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return False, None


def _validate_decimal(value: str) -> tuple[bool, Decimal | None]:
    """Validate and parse decimal value."""
    try:
        # Remove currency symbols and commas
        cleaned = re.sub(r"[$£€,]", "", value)
        return True, Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return False, None


def _apply_transform(value: str, transform: str | None) -> str:
    """Apply a transform to a value."""
    if not transform:
        return value

    if transform == "uppercase":
        return value.upper()
    elif transform == "lowercase":
        return value.lower()
    elif transform == "trim":
        return value.strip()
    elif transform.startswith("date:"):
        # Date format transform
        _valid, normalized = _validate_date(value)
        if normalized:
            return normalized
    return value


class ImportService:
    """Service for handling data imports."""

    async def get_template(self, entity_type: ImportEntityType) -> ImportTemplateResponse:
        """Get import template for an entity type."""
        fields = FIELD_DEFINITIONS.get(entity_type, [])
        csv_headers = [f.display_name for f in fields]
        example_rows = [
            {f.display_name: f.example_values[0] if f.example_values else "" for f in fields}
        ]

        return ImportTemplateResponse(
            entity_type=entity_type,
            fields=fields,
            example_rows=example_rows,
            csv_headers=csv_headers,
        )

    async def upload_file(
        self,
        content: bytes,
        filename: str,
        entity_type: ImportEntityType,
    ) -> ImportJobResponse:
        """Upload and parse an import file."""
        # Detect file type and parse
        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            columns, rows = _parse_csv(content)
        elif lower_name.endswith((".xlsx", ".xls")):
            columns, rows = _parse_excel(content)
        else:
            raise BadRequestException(
                "Unsupported file format. Please upload a CSV or Excel file."
            )

        if not rows:
            raise BadRequestException("File contains no data rows")

        # Auto-detect column mappings
        detected_mappings = _auto_detect_mappings(columns, entity_type)

        # Create import job
        import_id = str(uuid.uuid4())
        job_data = {
            "import_id": import_id,
            "entity_type": entity_type,
            "filename": filename,
            "status": ImportStatus.PENDING,
            "columns": columns,
            "raw_rows": rows,
            "detected_mappings": detected_mappings,
            "mappings": [],
            "default_values": {},
            "errors": [],
            "warnings": [],
            "preview_rows": [],
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        }
        _import_jobs[import_id] = job_data

        from typing import cast as _cast
        return ImportJobResponse(
            import_id=import_id,
            entity_type=entity_type,
            filename=filename,
            status=ImportStatus.PENDING,
            created_at=_cast(datetime, job_data["created_at"]),
            updated_at=_cast(datetime, job_data["updated_at"]),
            total_rows=len(rows),
            mappings=[],
        )

    async def map_columns(
        self,
        import_id: str,
        mappings: list[ColumnMapping],
        default_values: dict[str, Any] | None = None,
    ) -> ImportJobResponse:
        """Set column mappings for an import job."""
        job = _import_jobs.get(import_id)
        if not job:
            raise NotFoundException("Import job not found")

        job["mappings"] = [
            {
                "source_column": m.source_column,
                "target_field": m.target_field,
                "transform": m.transform,
            }
            for m in mappings
        ]
        job["default_values"] = default_values or {}
        job["status"] = ImportStatus.MAPPING
        job["updated_at"] = datetime.utcnow()

        return ImportJobResponse(
            import_id=import_id,
            entity_type=job["entity_type"],
            filename=job["filename"],
            status=ImportStatus.MAPPING,
            created_at=job["created_at"],
            updated_at=job["updated_at"],
            total_rows=len(job["raw_rows"]),
            mappings=mappings,
        )

    async def validate(  # noqa: PLR0912, PLR0915
        self,
        db: AsyncSession,
        import_id: str,
        skip_duplicates: bool = False,
    ) -> dict[str, Any]:
        """Validate import data and return preview."""
        job = _import_jobs.get(import_id)
        if not job:
            raise NotFoundException("Import job not found")

        job["status"] = ImportStatus.VALIDATING
        job["errors"] = []
        job["warnings"] = []
        job["updated_at"] = datetime.utcnow()

        entity_type = job["entity_type"]
        field_defs = FIELD_DEFINITIONS.get(entity_type, [])
        required_fields = {f.name for f in field_defs if f.required}
        field_types = {f.name: f for f in field_defs}

        # Build mapping lookup
        mapping_lookup: dict[str, str] = {}
        transform_lookup: dict[str, str | None] = {}
        for m in job["mappings"]:
            mapping_lookup[m["source_column"]] = m["target_field"]
            transform_lookup[m["source_column"]] = m.get("transform")

        preview_rows: list[dict[str, Any]] = []
        valid_count = 0
        invalid_count = 0
        warning_count = 0

        # Load reference data for validation
        existing_clients = await self._load_client_emails(db)
        existing_programs = await self._load_program_titles(db)
        existing_users = await self._load_user_emails(db)

        for row_num, raw_row in enumerate(job["raw_rows"], start=1):
            row_errors: list[dict[str, Any]] = []
            row_warnings: list[dict[str, Any]] = []
            mapped_data: dict[str, Any] = {}

            # Apply mappings and defaults
            for col in job["columns"]:
                target_field = mapping_lookup.get(col)
                if target_field:
                    value = str(raw_row.get(col, "")).strip()
                    if value:
                        transform = transform_lookup.get(col)
                        value = _apply_transform(value, transform)
                    mapped_data[target_field] = value

            # Apply defaults for missing fields
            for field_name, default_val in job.get("default_values", {}).items():
                if field_name not in mapped_data or not mapped_data[field_name]:
                    mapped_data[field_name] = default_val

            # Validate required fields
            for req_field in required_fields:
                value = mapped_data.get(req_field) or ""
                if not value:
                    row_errors.append({
                        "row_number": row_num,
                        "field": req_field,
                        "error_type": "required",
                        "message": f"{req_field.replace('_', ' ').title()} is required",
                        "value": None,
                    })

            # Validate field types and formats
            for field_name, value in mapped_data.items():
                if not value:
                    continue

                field_def = field_types.get(field_name)
                if not field_def:
                    continue

                # Email validation
                if field_def.field_type == "email":
                    if not _validate_email(value):
                        row_errors.append({
                            "row_number": row_num,
                            "field": field_name,
                            "error_type": "format",
                            "message": f"Invalid email format: {value}",
                            "value": value,
                        })

                # Phone validation
                elif field_def.field_type == "phone":
                    if not _validate_phone(value):
                        row_errors.append({
                            "row_number": row_num,
                            "field": field_name,
                            "error_type": "format",
                            "message": f"Invalid phone format: {value}",
                            "value": value,
                        })

                # Date validation
                elif field_def.field_type == "date":
                    valid, normalized = _validate_date(value)
                    if not valid:
                        row_errors.append({
                            "row_number": row_num,
                            "field": field_name,
                            "error_type": "format",
                            "message": f"Invalid date format: {value}. Use YYYY-MM-DD",
                            "value": value,
                        })
                    else:
                        mapped_data[field_name] = normalized

                # Number validation
                elif field_def.field_type == "number":
                    valid, parsed = _validate_decimal(value)
                    if not valid:
                        row_errors.append({
                            "row_number": row_num,
                            "field": field_name,
                            "error_type": "format",
                            "message": f"Invalid number format: {value}",
                            "value": value,
                        })
                    else:
                        mapped_data[field_name] = str(parsed)

                # Enum validation
                elif (
                    field_def.field_type == "enum"
                    and field_def.enum_values
                    and value.lower() not in [v.lower() for v in field_def.enum_values]
                ):
                    row_errors.append({
                        "row_number": row_num,
                        "field": field_name,
                        "error_type": "value",
                        "message": (
                            f"Invalid value: {value}. "
                            f"Must be one of: {', '.join(field_def.enum_values)}"
                        ),
                        "value": value,
                    })

            # Reference validation (entity-specific)
            if entity_type == ImportEntityType.CLIENTS:
                # Check for duplicates
                has_name = not skip_duplicates and mapped_data.get("legal_name")
                if has_name or mapped_data.get("primary_email"):
                    duplicates = await check_duplicates(
                        db,
                        legal_name=mapped_data.get("legal_name"),
                        primary_email=mapped_data.get("primary_email"),
                        phone=mapped_data.get("phone"),
                    )
                    for dup in duplicates:
                        row_warnings.append({
                            "row_number": row_num,
                            "field": "legal_name",
                            "warning_type": "duplicate_match",
                            "message": (
                                f"Potential duplicate: {dup.legal_name} "
                                f"({int(dup.similarity_score * 100)}% match)"
                            ),
                            "value": mapped_data.get("legal_name"),
                            "existing_id": dup.client_id,
                            "existing_name": dup.legal_name,
                        })

                # Validate RM email
                rm_email = mapped_data.get("assigned_rm_email")
                if rm_email and rm_email not in existing_users:
                    row_errors.append({
                        "row_number": row_num,
                        "field": "assigned_rm_email",
                        "error_type": "reference",
                        "message": f"Relationship manager not found: {rm_email}",
                        "value": rm_email,
                    })

            elif entity_type == ImportEntityType.PROGRAMS:
                # Validate client exists
                client_email = mapped_data.get("client_email")
                if client_email and client_email not in existing_clients:
                    row_errors.append({
                        "row_number": row_num,
                        "field": "client_email",
                        "error_type": "reference",
                        "message": f"Client not found with email: {client_email}",
                        "value": client_email,
                    })

            elif entity_type == ImportEntityType.TASKS:
                # Validate program exists
                program_title = mapped_data.get("program_title")
                if program_title and program_title not in existing_programs:
                    row_warnings.append({
                        "row_number": row_num,
                        "field": "program_title",
                        "warning_type": "reference",
                        "message": (
                            f"Program not found: {program_title}. "
                            "Task will be created without program association."
                        ),
                        "value": program_title,
                    })

                # Validate assignee
                assignee_email = mapped_data.get("assigned_to_email")
                if assignee_email and assignee_email not in existing_users:
                    row_errors.append({
                        "row_number": row_num,
                        "field": "assigned_to_email",
                        "error_type": "reference",
                        "message": f"User not found: {assignee_email}",
                        "value": assignee_email,
                    })

            # Track row status
            is_valid = len(row_errors) == 0
            if is_valid:
                valid_count += 1
            else:
                invalid_count += 1
            if row_warnings:
                warning_count += 1

            job["errors"].extend(row_errors)
            job["warnings"].extend(row_warnings)

            # Add to preview (limit to first 100 rows)
            if len(preview_rows) < 100:
                preview_rows.append({
                    "row_number": row_num,
                    "data": raw_row,
                    "mapped_data": mapped_data,
                    "is_valid": is_valid,
                    "errors": row_errors,
                    "warnings": row_warnings,
                })

        job["preview_rows"] = preview_rows
        job["status"] = ImportStatus.PREVIEW
        job["valid_rows"] = valid_count
        job["invalid_rows"] = invalid_count
        job["rows_with_warnings"] = warning_count
        job["updated_at"] = datetime.utcnow()

        return {
            "import_id": import_id,
            "status": job["status"],
            "total_rows": len(job["raw_rows"]),
            "valid_rows": valid_count,
            "invalid_rows": invalid_count,
            "rows_with_warnings": warning_count,
            "errors": job["errors"],
            "warnings": job["warnings"],
            "preview_rows": preview_rows,
        }

    async def confirm_import(  # noqa: PLR0912, PLR0915
        self,
        db: AsyncSession,
        import_id: str,
        skip_invalid_rows: bool = True,
        skip_warnings: bool = False,
        created_by_id: uuid.UUID | None = None,
    ) -> dict[str, Any]:
        """Execute the import after confirmation."""
        job = _import_jobs.get(import_id)
        if not job:
            raise NotFoundException("Import job not found")

        if job["status"] != ImportStatus.PREVIEW:
            raise BadRequestException("Import must be validated before confirmation")

        job["status"] = ImportStatus.IMPORTING
        job["updated_at"] = datetime.utcnow()

        entity_type = job["entity_type"]
        created_ids: list[uuid.UUID] = []
        import_errors: list[dict[str, Any]] = []
        imported_count = 0
        skipped_count = 0
        failed_count = 0

        # Build error row lookup
        error_rows: set[int] = set()
        warning_rows: set[int] = set()
        for err in job["errors"]:
            error_rows.add(err["row_number"])
        for warn in job["warnings"]:
            warning_rows.add(warn["row_number"])

        # Load reference data for lookups
        client_email_to_id = await self._load_client_email_to_id(db)
        user_email_to_id = await self._load_user_email_to_id(db)

        # Suppress per-row audit log entries during bulk import to avoid
        # write amplification (one audit entry per inserted row).
        db.info["skip_audit"] = True

        for preview_row in job["preview_rows"]:
            row_num = preview_row["row_number"]
            mapped_data = preview_row["mapped_data"]

            # Skip invalid rows if requested
            if skip_invalid_rows and row_num in error_rows:
                skipped_count += 1
                continue

            # Skip rows with warnings if requested
            if skip_warnings and row_num in warning_rows:
                skipped_count += 1
                continue

            try:
                if entity_type == ImportEntityType.CLIENTS:
                    created_id = await self._import_client(
                        db, mapped_data, user_email_to_id, created_by_id
                    )
                    if created_id:
                        created_ids.append(created_id)
                        imported_count += 1

                elif entity_type == ImportEntityType.PARTNERS:
                    created_id = await self._import_partner(db, mapped_data, created_by_id)
                    if created_id:
                        created_ids.append(created_id)
                        imported_count += 1

                elif entity_type == ImportEntityType.PROGRAMS:
                    created_id = await self._import_program(
                        db, mapped_data, client_email_to_id, created_by_id
                    )
                    if created_id:
                        created_ids.append(created_id)
                        imported_count += 1

                elif entity_type == ImportEntityType.TASKS:
                    # Tasks import is more complex - needs program/milestone context
                    # For now, create as standalone tasks
                    created_id = await self._import_task(db, mapped_data, user_email_to_id)
                    if created_id:
                        created_ids.append(created_id)
                        imported_count += 1

            except Exception as e:
                logger.error(f"Import failed for row {row_num}: {e}")
                import_errors.append({
                    "row_number": row_num,
                    "error_type": "import",
                    "message": str(e),
                    "field": None,
                    "value": None,
                })
                failed_count += 1

        job["status"] = ImportStatus.COMPLETED if failed_count == 0 else ImportStatus.FAILED
        job["imported_rows"] = imported_count
        job["skipped_rows"] = skipped_count
        job["failed_rows"] = failed_count
        job["created_ids"] = [str(id) for id in created_ids]
        job["errors"].extend(import_errors)
        job["updated_at"] = datetime.utcnow()

        return {
            "import_id": import_id,
            "status": job["status"],
            "total_rows": len(job["raw_rows"]),
            "imported_rows": imported_count,
            "skipped_rows": skipped_count,
            "failed_rows": failed_count,
            "created_ids": created_ids,
            "errors": import_errors,
        }

    async def get_error_report(self, import_id: str) -> dict[str, Any]:
        """Generate a downloadable error report CSV."""
        job = _import_jobs.get(import_id)
        if not job:
            raise NotFoundException("Import job not found")

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(["Row", "Column", "Field", "Error Type", "Message", "Value"])

        # Write errors
        for err in job.get("errors", []):
            writer.writerow([
                err.get("row_number", ""),
                err.get("column", ""),
                err.get("field", ""),
                err.get("error_type", ""),
                err.get("message", ""),
                err.get("value", ""),
            ])

        # Write warnings
        writer.writerow([])  # Blank row
        writer.writerow(["Warnings:"])
        for warn in job.get("warnings", []):
            writer.writerow([
                warn.get("row_number", ""),
                warn.get("column", ""),
                warn.get("field", ""),
                warn.get("warning_type", ""),
                warn.get("message", ""),
                warn.get("value", ""),
            ])

        content = output.getvalue()
        encoded = base64.b64encode(content.encode()).decode()

        return {
            "import_id": import_id,
            "filename": f"import_errors_{import_id[:8]}.csv",
            "content_type": "text/csv",
            "content": encoded,
        }

    async def get_job(self, import_id: str) -> ImportJobResponse | None:
        """Get import job by ID."""
        job = _import_jobs.get(import_id)
        if not job:
            return None

        mappings = [
            ColumnMapping(
                source_column=m["source_column"],
                target_field=m["target_field"],
                transform=m.get("transform"),
            )
            for m in job.get("mappings", [])
        ]

        return ImportJobResponse(
            import_id=import_id,
            entity_type=job["entity_type"],
            filename=job["filename"],
            status=job["status"],
            created_at=job["created_at"],
            updated_at=job["updated_at"],
            total_rows=len(job.get("raw_rows", [])),
            valid_rows=job.get("valid_rows"),
            invalid_rows=job.get("invalid_rows"),
            imported_rows=job.get("imported_rows"),
            errors=[ImportError(**e) for e in job.get("errors", [])],
            warnings=[ImportWarning(**w) for w in job.get("warnings", [])],
            mappings=mappings,
        )

    async def list_jobs(self, limit: int = 20) -> list[ImportJobResponse]:
        """List recent import jobs."""
        jobs = sorted(
            _import_jobs.values(),
            key=lambda j: j["created_at"],
            reverse=True,
        )[:limit]

        results = []
        for j in jobs:
            job = await self.get_job(j["import_id"])
            if job:
                results.append(job)
        return results

    # --- Helper methods ---

    async def _load_client_emails(self, db: AsyncSession) -> set[str]:
        """Load all client emails for reference validation."""
        result = await db.execute(select(ClientProfile.primary_email))
        return {row[0].lower() for row in result.fetchall()}

    async def _load_partner_emails(self, db: AsyncSession) -> set[str]:
        """Load all partner emails for reference validation."""
        result = await db.execute(select(PartnerProfile.contact_email))
        return {row[0].lower() for row in result.fetchall()}

    async def _load_program_titles(self, db: AsyncSession) -> set[str]:
        """Load all program titles for reference validation."""
        result = await db.execute(select(Program.title))
        return {row[0] for row in result.fetchall()}

    async def _load_user_emails(self, db: AsyncSession) -> set[str]:
        """Load all user emails for reference validation."""
        result = await db.execute(select(User.email))
        return {row[0].lower() for row in result.fetchall()}

    async def _load_client_email_to_id(self, db: AsyncSession) -> dict[str, uuid.UUID]:
        """Load client email to ID mapping."""
        result = await db.execute(select(ClientProfile.id, ClientProfile.primary_email))
        return {row[1].lower(): row[0] for row in result.fetchall()}

    async def _load_user_email_to_id(self, db: AsyncSession) -> dict[str, uuid.UUID]:
        """Load user email to ID mapping."""
        result = await db.execute(select(User.id, User.email))
        return {row[1].lower(): row[0] for row in result.fetchall()}

    async def _import_client(
        self,
        db: AsyncSession,
        data: dict[str, Any],
        user_email_to_id: dict[str, uuid.UUID],
        created_by_id: uuid.UUID | None,
    ) -> uuid.UUID | None:
        """Import a single client."""
        from app.schemas.client_profile import ClientProfileCreate

        # Map assigned RM email to ID
        assigned_rm_id = None
        rm_email = data.get("assigned_rm_email")
        if rm_email:
            assigned_rm_id = user_email_to_id.get(rm_email.lower())

        create_data = ClientProfileCreate(
            legal_name=data.get("legal_name", ""),
            display_name=data.get("display_name") or None,
            entity_type=data.get("entity_type") or None,
            jurisdiction=data.get("jurisdiction") or None,
            tax_id=data.get("tax_id") or None,
            primary_email=data.get("primary_email", ""),
            secondary_email=data.get("secondary_email") or None,
            phone=data.get("phone") or None,
            address=data.get("address") or None,
            communication_preference=data.get("communication_preference") or None,
            sensitivities=data.get("sensitivities") or None,
            special_instructions=data.get("special_instructions") or None,
        )

        client = await client_service.create_intake(
            db,
            data=create_data,
            created_by_id=created_by_id or uuid.UUID("00000000-0000-0000-0000-000000000000"),
        )

        # Update assigned RM if provided
        if assigned_rm_id and client:
            client.assigned_rm_id = assigned_rm_id
            await db.commit()

        return client.id if client else None

    async def _import_partner(
        self,
        db: AsyncSession,
        data: dict[str, Any],
        created_by_id: uuid.UUID | None,
    ) -> uuid.UUID | None:
        """Import a single partner."""
        from app.models.enums import PartnerStatus
        from app.models.partner import PartnerProfile

        capabilities = []
        if data.get("capabilities"):
            capabilities = [c.strip() for c in data["capabilities"].split(",")]

        geographies = []
        if data.get("geographies"):
            geographies = [g.strip() for g in data["geographies"].split(",")]

        partner = PartnerProfile(
            firm_name=data.get("firm_name", ""),
            contact_name=data.get("contact_name", ""),
            contact_email=data.get("contact_email", ""),
            contact_phone=data.get("contact_phone") or None,
            capabilities=capabilities,
            geographies=geographies,
            notes=data.get("notes") or None,
            status=PartnerStatus.active.value,
            availability_status="available",
            compliance_verified=False,
            total_assignments=0,
            completed_assignments=0,
            created_by=created_by_id or uuid.UUID("00000000-0000-0000-0000-000000000000"),
        )

        db.add(partner)
        await db.commit()
        await db.refresh(partner)
        return partner.id

    async def _import_program(
        self,
        db: AsyncSession,
        data: dict[str, Any],
        client_email_to_id: dict[str, uuid.UUID],
        created_by_id: uuid.UUID | None,
    ) -> uuid.UUID | None:
        """Import a single program."""
        from app.models.enums import ProgramStatus
        from app.models.program import Program

        client_email = data.get("client_email", "").lower()
        client_id = client_email_to_id.get(client_email)
        if not client_id:
            raise ValueError(f"Client not found: {client_email}")

        budget = None
        if data.get("budget_envelope"):
            with contextlib.suppress(InvalidOperation, ValueError):
                budget = Decimal(data["budget_envelope"])

        status_val = data.get("status") or ProgramStatus.intake.value

        program = Program(
            client_id=client_id,
            title=data.get("title", ""),
            objectives=data.get("objectives") or None,
            scope=data.get("scope") or None,
            budget_envelope=budget,
            start_date=data.get("start_date") or None,
            end_date=data.get("end_date") or None,
            status=status_val,
            created_by=created_by_id or uuid.UUID("00000000-0000-0000-0000-000000000000"),
        )

        db.add(program)
        await db.commit()
        await db.refresh(program)
        return program.id

    async def _import_task(
        self,
        db: AsyncSession,
        data: dict[str, Any],
        user_email_to_id: dict[str, uuid.UUID],
    ) -> uuid.UUID | None:
        """Import a single task (as a standalone task)."""
        # Note: Full task import requires program/milestone context
        # This creates standalone tasks which can be linked later
        from app.models.task import Task

        assignee_id = None
        assignee_email = data.get("assigned_to_email")
        if assignee_email:
            assignee_id = user_email_to_id.get(assignee_email.lower())

        priority = TaskPriority.medium
        if data.get("priority"):
            with contextlib.suppress(ValueError):
                priority = TaskPriority(data["priority"].lower())

        status_val = TaskStatus.todo
        if data.get("status"):
            with contextlib.suppress(ValueError):
                status_val = TaskStatus(data["status"].lower())

        task = Task(
            title=data.get("title", ""),
            description=data.get("description") or None,
            due_date=data.get("due_date") or None,
            assigned_to=assignee_id,
            priority=priority,
            status=status_val,
            # Note: milestone_id is required but not available for standalone tasks
            # This would need to be handled in a full implementation
        )

        db.add(task)
        await db.commit()
        await db.refresh(task)
        return task.id


import_service = ImportService()
