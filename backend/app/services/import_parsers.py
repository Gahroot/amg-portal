"""Pure CSV/Excel parsing, column detection, field mapping, and row validation.

No database imports or SQLAlchemy — only stdlib and openpyxl.
"""

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from email.utils import parseaddr
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import BadRequestException
from app.schemas.import_schemas import (
    ImportEntityType,
    ImportFieldDefinition,
)

# ---------------------------------------------------------------------------
# Field definitions
# ---------------------------------------------------------------------------

CLIENT_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="legal_name",
        display_name="Legal Name",
        description="Client's legal name",
        required=True,
        field_type="string",
        example_values=["John Smith", "Acme Corporation"],
    ),
    ImportFieldDefinition(
        name="display_name",
        display_name="Display Name",
        description="Preferred name or nickname",
        required=False,
        field_type="string",
        example_values=["John", "Acme Corp"],
    ),
    ImportFieldDefinition(
        name="entity_type",
        display_name="Entity Type",
        description="Type of entity",
        required=False,
        field_type="enum",
        enum_values=["individual", "family_office", "trust", "corporation"],
        example_values=["individual", "corporation"],
    ),
    ImportFieldDefinition(
        name="jurisdiction",
        display_name="Jurisdiction",
        description="Primary jurisdiction",
        required=False,
        field_type="string",
        example_values=["United States", "United Kingdom", "Singapore"],
    ),
    ImportFieldDefinition(
        name="tax_id",
        display_name="Tax ID",
        description="Tax identification number",
        required=False,
        field_type="string",
        example_values=["123-45-6789"],
    ),
    ImportFieldDefinition(
        name="primary_email",
        display_name="Primary Email",
        description="Primary contact email",
        required=True,
        field_type="email",
        example_values=["john@example.com"],
    ),
    ImportFieldDefinition(
        name="secondary_email",
        display_name="Secondary Email",
        description="Secondary contact email",
        required=False,
        field_type="email",
        example_values=["john.d@example.com"],
    ),
    ImportFieldDefinition(
        name="phone",
        display_name="Phone",
        description="Contact phone number",
        required=False,
        field_type="phone",
        example_values=["+1 (555) 123-4567"],
    ),
    ImportFieldDefinition(
        name="address",
        display_name="Address",
        description="Physical address",
        required=False,
        field_type="string",
        example_values=["123 Main St, New York, NY 10001"],
    ),
    ImportFieldDefinition(
        name="communication_preference",
        display_name="Communication Preference",
        description="Preferred communication channel",
        required=False,
        field_type="enum",
        enum_values=["email", "phone", "sms", "whatsapp"],
        example_values=["email"],
    ),
    ImportFieldDefinition(
        name="sensitivities",
        display_name="Sensitivities",
        description="Any sensitivities to be aware of",
        required=False,
        field_type="string",
        example_values=["Prefers evening calls"],
    ),
    ImportFieldDefinition(
        name="special_instructions",
        display_name="Special Instructions",
        description="Special handling instructions",
        required=False,
        field_type="string",
        example_values=["Contact assistant first"],
    ),
    ImportFieldDefinition(
        name="birth_date",
        display_name="Birth Date",
        description="Date of birth (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["1980-01-15"],
    ),
    ImportFieldDefinition(
        name="assigned_rm_email",
        display_name="Assigned RM Email",
        description="Email of the relationship manager to assign",
        required=False,
        field_type="email",
        example_values=["rm@example.com"],
    ),
]

PARTNER_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="firm_name",
        display_name="Firm Name",
        description="Name of the partner firm",
        required=True,
        field_type="string",
        example_values=["Global Advisory Partners"],
    ),
    ImportFieldDefinition(
        name="contact_name",
        display_name="Contact Name",
        description="Primary contact person's name",
        required=True,
        field_type="string",
        example_values=["Jane Doe"],
    ),
    ImportFieldDefinition(
        name="contact_email",
        display_name="Contact Email",
        description="Primary contact email",
        required=True,
        field_type="email",
        example_values=["jane@globaladvisory.com"],
    ),
    ImportFieldDefinition(
        name="contact_phone",
        display_name="Contact Phone",
        description="Contact phone number",
        required=False,
        field_type="phone",
        example_values=["+44 20 1234 5678"],
    ),
    ImportFieldDefinition(
        name="capabilities",
        display_name="Capabilities",
        description="Comma-separated list of capabilities",
        required=False,
        field_type="string",
        example_values=["Legal, Tax Advisory, Immigration"],
    ),
    ImportFieldDefinition(
        name="geographies",
        display_name="Geographies",
        description="Comma-separated list of geographies served",
        required=False,
        field_type="string",
        example_values=["United States, Europe, Asia"],
    ),
    ImportFieldDefinition(
        name="notes",
        display_name="Notes",
        description="Additional notes about the partner",
        required=False,
        field_type="string",
        example_values=["Preferred partner for tax matters"],
    ),
]

PROGRAM_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="title",
        display_name="Program Title",
        description="Title of the program",
        required=True,
        field_type="string",
        example_values=["Estate Planning 2024"],
    ),
    ImportFieldDefinition(
        name="client_email",
        display_name="Client Email",
        description="Email of the client (must exist)",
        required=True,
        field_type="email",
        example_values=["client@example.com"],
    ),
    ImportFieldDefinition(
        name="objectives",
        display_name="Objectives",
        description="Program objectives",
        required=False,
        field_type="string",
        example_values=["Set up trust structure"],
    ),
    ImportFieldDefinition(
        name="scope",
        display_name="Scope",
        description="Program scope",
        required=False,
        field_type="string",
        example_values=["US and UK jurisdictions"],
    ),
    ImportFieldDefinition(
        name="budget_envelope",
        display_name="Budget Envelope",
        description="Budget amount",
        required=False,
        field_type="number",
        example_values=["50000", "100000.00"],
    ),
    ImportFieldDefinition(
        name="start_date",
        display_name="Start Date",
        description="Program start date (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["2024-01-01"],
    ),
    ImportFieldDefinition(
        name="end_date",
        display_name="End Date",
        description="Program end date (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["2024-12-31"],
    ),
    ImportFieldDefinition(
        name="status",
        display_name="Status",
        description="Program status",
        required=False,
        field_type="enum",
        enum_values=["planning", "active", "on_hold", "completed", "cancelled"],
        example_values=["planning", "active"],
    ),
]

TASK_FIELDS: list[ImportFieldDefinition] = [
    ImportFieldDefinition(
        name="title",
        display_name="Task Title",
        description="Title of the task",
        required=True,
        field_type="string",
        example_values=["Review legal documents"],
    ),
    ImportFieldDefinition(
        name="program_title",
        display_name="Program Title",
        description="Title of the associated program",
        required=False,
        field_type="string",
        example_values=["Estate Planning 2024"],
    ),
    ImportFieldDefinition(
        name="milestone_title",
        display_name="Milestone Title",
        description="Title of the associated milestone",
        required=False,
        field_type="string",
        example_values=["Phase 1: Research"],
    ),
    ImportFieldDefinition(
        name="description",
        display_name="Description",
        description="Task description",
        required=False,
        field_type="string",
        example_values=["Review all legal documents for compliance"],
    ),
    ImportFieldDefinition(
        name="due_date",
        display_name="Due Date",
        description="Task due date (YYYY-MM-DD)",
        required=False,
        field_type="date",
        example_values=["2024-03-15"],
    ),
    ImportFieldDefinition(
        name="assigned_to_email",
        display_name="Assigned To Email",
        description="Email of the user to assign",
        required=False,
        field_type="email",
        example_values=["coordinator@example.com"],
    ),
    ImportFieldDefinition(
        name="priority",
        display_name="Priority",
        description="Task priority",
        required=False,
        field_type="enum",
        enum_values=["low", "medium", "high", "urgent"],
        example_values=["medium", "high"],
    ),
    ImportFieldDefinition(
        name="status",
        display_name="Status",
        description="Task status",
        required=False,
        field_type="enum",
        enum_values=["pending", "in_progress", "completed", "blocked"],
        example_values=["pending", "in_progress"],
    ),
]

FIELD_DEFINITIONS: dict[ImportEntityType, list[ImportFieldDefinition]] = {
    ImportEntityType.CLIENTS: CLIENT_FIELDS,
    ImportEntityType.PARTNERS: PARTNER_FIELDS,
    ImportEntityType.PROGRAMS: PROGRAM_FIELDS,
    ImportEntityType.TASKS: TASK_FIELDS,
}

# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------


def _normalize_column_name(name: str) -> str:
    """Normalize column name for matching."""
    normalized = name.lower().strip()
    normalized = re.sub(r"[\s\-]+", "_", normalized)
    normalized = re.sub(r"[^a-z0-9_]", "", normalized)
    return normalized


def auto_detect_mappings(columns: list[str], entity_type: ImportEntityType) -> dict[str, str]:
    """Auto-detect column-to-field mappings based on column names."""
    field_defs = FIELD_DEFINITIONS.get(entity_type, [])

    field_variations: dict[str, str] = {}
    for field in field_defs:
        field_variations[_normalize_column_name(field.name)] = field.name
        field_variations[_normalize_column_name(field.display_name)] = field.name
        if field.name == "primary_email":
            field_variations["email"] = field.name
            field_variations["email_address"] = field.name
        if field.name == "legal_name":
            field_variations["name"] = field.name
            field_variations["client_name"] = field.name
        if field.name == "firm_name":
            field_variations["name"] = field.name
            field_variations["partner_name"] = field.name
        if field.name == "contact_email":
            field_variations["email"] = field.name
        if field.name == "contact_name":
            field_variations["name"] = field.name

    mappings: dict[str, str] = {}
    for col in columns:
        normalized = _normalize_column_name(col)
        if normalized in field_variations:
            mappings[col] = field_variations[normalized]

    return mappings


# ---------------------------------------------------------------------------
# File parsers
# ---------------------------------------------------------------------------


def parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse CSV content and return (headers, rows)."""
    text = content.decode("utf-8-sig")  # Handle BOM
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise BadRequestException("CSV file has no headers")
    columns = list(reader.fieldnames)
    rows = list(reader)
    return columns, rows


def parse_excel(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse Excel content and return (headers, rows)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise BadRequestException("Excel file has no sheets")

    headers: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        value = cell.value
        headers.append(str(value) if value is not None else f"Column_{cell.column_letter}")

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict: dict[str, str] = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            row_dict[header] = str(value) if value is not None else ""
        rows.append(row_dict)

    wb.close()
    return headers, rows


# ---------------------------------------------------------------------------
# Field validators
# ---------------------------------------------------------------------------


def validate_email(value: str) -> bool:
    """Validate email format."""
    _, email = parseaddr(value)
    return bool(email and "@" in email and "." in email.split("@")[1])


def validate_phone(value: str) -> bool:
    """Validate phone format (basic check)."""
    cleaned = re.sub(r"[\d\s\-\(\)\+]", "", value)
    return len(cleaned) == 0 and len(re.sub(r"\D", "", value)) >= 7


def validate_date(value: str) -> tuple[bool, str | None]:
    """Validate and normalize date format. Returns (valid, normalized_iso_string)."""
    from datetime import datetime as dt

    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"]
    for fmt in formats:
        try:
            parsed = dt.strptime(value, fmt)
            return True, parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return False, None


def validate_decimal(value: str) -> tuple[bool, Decimal | None]:
    """Validate and parse decimal value."""
    try:
        cleaned = re.sub(r"[$£€,]", "", value)
        return True, Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return False, None


# ---------------------------------------------------------------------------
# Transform helpers
# ---------------------------------------------------------------------------


def apply_transform(value: str, transform: str | None) -> str:
    """Apply a named transform to a string value."""
    if not transform:
        return value
    if transform == "uppercase":
        return value.upper()
    if transform == "lowercase":
        return value.lower()
    if transform == "trim":
        return value.strip()
    if transform.startswith("date:"):
        _valid, normalized = validate_date(value)
        if normalized:
            return normalized
    return value


# ---------------------------------------------------------------------------
# Row validation
# ---------------------------------------------------------------------------


def validate_row(  # noqa: PLR0912
    row_num: int,
    raw_row: dict[str, str],
    columns: list[str],
    mapping_lookup: dict[str, str],
    transform_lookup: dict[str, str | None],
    default_values: dict[str, Any],
    field_defs: list[ImportFieldDefinition],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """Apply mappings, defaults, and format validation to a single row.

    Returns (mapped_data, row_errors, row_warnings).
    Note: reference/duplicate checks are performed in import_service.py where
    DB access is available.
    """
    required_fields = {f.name for f in field_defs if f.required}
    field_types = {f.name: f for f in field_defs}

    mapped_data: dict[str, Any] = {}
    row_errors: list[dict[str, Any]] = []
    row_warnings: list[dict[str, Any]] = []

    # Apply column mappings
    for col in columns:
        target_field = mapping_lookup.get(col)
        if target_field:
            value = str(raw_row.get(col, "")).strip()
            if value:
                transform = transform_lookup.get(col)
                value = apply_transform(value, transform)
            mapped_data[target_field] = value

    # Apply defaults for missing fields
    for field_name, default_val in default_values.items():
        if field_name not in mapped_data or not mapped_data[field_name]:
            mapped_data[field_name] = default_val

    # Required field validation
    for req_field in required_fields:
        if not mapped_data.get(req_field):
            row_errors.append(
                {
                    "row_number": row_num,
                    "field": req_field,
                    "error_type": "required",
                    "message": f"{req_field.replace('_', ' ').title()} is required",
                    "value": None,
                }
            )

    # Format/type validation
    for field_name, value in mapped_data.items():
        if not value:
            continue
        field_def = field_types.get(field_name)
        if not field_def:
            continue

        if field_def.field_type == "email":
            if not validate_email(value):
                row_errors.append(
                    {
                        "row_number": row_num,
                        "field": field_name,
                        "error_type": "format",
                        "message": f"Invalid email format: {value}",
                        "value": value,
                    }
                )

        elif field_def.field_type == "phone":
            if not validate_phone(value):
                row_errors.append(
                    {
                        "row_number": row_num,
                        "field": field_name,
                        "error_type": "format",
                        "message": f"Invalid phone format: {value}",
                        "value": value,
                    }
                )

        elif field_def.field_type == "date":
            valid, normalized = validate_date(value)
            if not valid:
                row_errors.append(
                    {
                        "row_number": row_num,
                        "field": field_name,
                        "error_type": "format",
                        "message": f"Invalid date format: {value}. Use YYYY-MM-DD",
                        "value": value,
                    }
                )
            else:
                mapped_data[field_name] = normalized

        elif field_def.field_type == "number":
            valid, parsed = validate_decimal(value)
            if not valid:
                row_errors.append(
                    {
                        "row_number": row_num,
                        "field": field_name,
                        "error_type": "format",
                        "message": f"Invalid number format: {value}",
                        "value": value,
                    }
                )
            else:
                mapped_data[field_name] = str(parsed)

        elif (
            field_def.field_type == "enum"
            and field_def.enum_values
            and value.lower() not in [v.lower() for v in field_def.enum_values]
        ):
            row_errors.append(
                {
                    "row_number": row_num,
                    "field": field_name,
                    "error_type": "value",
                    "message": (
                        f"Invalid value: {value}. "
                        f"Must be one of: {', '.join(field_def.enum_values)}"
                    ),
                    "value": value,
                }
            )

    return mapped_data, row_errors, row_warnings
